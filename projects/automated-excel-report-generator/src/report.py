import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Set matplotlib backend to Agg (headless) before importing pyplot
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

def generate_report(raw_df: pd.DataFrame, cleaned_df: pd.DataFrame, analysis: dict, output_path: str):
    """
    Orchestrates the creation of the styled Excel workbook.
    """
    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    # Create the workbook
    wb = openpyxl.Workbook()
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 1. Sheet 1: Raw Data
    ws_raw = wb.create_sheet(title="Raw Data")
    write_dataframe_sheet(ws_raw, raw_df)
    
    # 2. Sheet 2: Cleaned Data
    ws_cleaned = wb.create_sheet(title="Cleaned Data")
    write_dataframe_sheet(ws_cleaned, cleaned_df, date_format="YYYY-MM-DD")
    
    # 3. Sheet 3: Summary Report
    ws_summary = wb.create_sheet(title="Summary Report")
    write_summary_sheet(ws_summary, analysis)
    
    # 4. Generate Matplotlib Charts
    chart_paths = generate_charts_images(analysis, output_dir)
    
    # 5. Sheet 4: Charts
    ws_charts = wb.create_sheet(title="Charts")
    ws_charts.views.sheetView[0].showGridLines = True
    
    # Insert charts into sheet
    from openpyxl.drawing.image import Image
    
    # Title for Charts Sheet
    ws_charts.cell(row=2, column=2, value="Sales Analytics Dashboard").font = Font(name="Segoe UI", size=18, bold=True, color="1F4E78")
    
    # Anchor positions
    ws_charts.add_image(Image(chart_paths['top_products']), 'B4')
    ws_charts.add_image(Image(chart_paths['monthly_trend']), 'B25')
    ws_charts.add_image(Image(chart_paths['regional_share']), 'L4')
    
    # Save the workbook
    wb.save(output_path)
    return output_path

def write_dataframe_sheet(ws, df: pd.DataFrame, date_format=None):
    """
    Writes a dataframe to a worksheet and applies headers/column formatting.
    """
    ws.views.sheetView[0].showGridLines = True
    
    # Write header
    headers = list(df.columns)
    ws.append(headers)
    
    # Write rows
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
        
    # Styles
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Style header row (row 1)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Style data rows
    for row_idx in range(2, ws.max_row + 1):
        for col_idx, col_name in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            
            # Format and align cells based on dtype/name
            val = cell.value
            if isinstance(val, (int, float)):
                cell.alignment = right_align
                if col_name in ['Revenue', 'Unit Price', 'Amount', 'Sales']:
                    cell.number_format = '$#,##0.00'
                elif col_name in ['Quantity', 'Qty', 'Orders']:
                    cell.number_format = '#,##0'
            elif pd.api.types.is_datetime64_any_dtype(df[col_name]):
                cell.alignment = center_align
                if date_format:
                    cell.number_format = date_format
            elif col_name in ['Order ID', 'ID']:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is not None:
                # Custom formatting len check
                val_str = str(val)
                if isinstance(val, float) and cell.number_format == '$#,##0.00':
                    val_str = f"${val:,.2f}"
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def write_summary_sheet(ws, analysis: dict):
    """
    Creates a highly styled summary dashboard sheet with KPIs and tables.
    """
    ws.views.sheetView[0].showGridLines = True
    
    # Setup styles
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1F4E78")
    section_font = Font(name="Segoe UI", size=12, bold=True, color="2C5282")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    
    kpi_title_font = Font(name="Segoe UI", size=9, color="555555")
    kpi_value_font = Font(name="Segoe UI", size=16, bold=True, color="1F4E78")
    kpi_fill = PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )
    
    double_bottom_border = Border(
        bottom=Side(style='double', color='1F4E78'),
        top=Side(style='thin', color='CBD5E0')
    )

    # 1. Sheet Title
    ws.cell(row=2, column=2, value="Sales Summary Dashboard").font = title_font
    
    # 2. KPI Section (Rows 4-5)
    kpi_cols = [
        ("Total Revenue", analysis['kpis']['total_revenue'], "$#,##0.00", 2),
        ("Total Orders", analysis['kpis']['total_orders'], "#,##0", 4),
        ("Average Order Value", analysis['kpis']['average_order_value'], "$#,##0.00", 6)
    ]
    
    for title, val, num_fmt, col_idx in kpi_cols:
        # Merge cell blocks for card style
        ws.merge_cells(start_row=4, start_column=col_idx, end_row=4, end_column=col_idx+1)
        ws.merge_cells(start_row=5, start_column=col_idx, end_row=5, end_column=col_idx+1)
        
        # Title Cell
        c_title = ws.cell(row=4, column=col_idx, value=title)
        c_title.font = kpi_title_font
        c_title.alignment = Alignment(horizontal="center", vertical="center")
        
        # Value Cell
        c_val = ws.cell(row=5, column=col_idx, value=val)
        c_val.font = kpi_value_font
        c_val.number_format = num_fmt
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        
        # Style all cells in the block for background & borders
        for r in [4, 5]:
            for c in [col_idx, col_idx+1]:
                cell = ws.cell(row=r, column=c)
                cell.fill = kpi_fill
                cell.border = thin_border

    # Helper function to write tables
    def write_table(start_row, start_col, title, df, number_formats):
        # Section title
        ws.cell(row=start_row, column=start_col, value=title).font = section_font
        
        # Table headers
        headers = list(df.columns)
        for col_offset, h in enumerate(headers):
            cell = ws.cell(row=start_row+1, column=start_col+col_offset, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Table data
        for row_offset, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
            r_idx = start_row + 2 + row_offset
            for col_offset, val in enumerate(row):
                c_idx = start_col + col_offset
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(name="Segoe UI", size=9)
                cell.border = thin_border
                
                # Align and format
                col_name = headers[col_offset]
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if col_name in number_formats:
                        cell.number_format = number_formats[col_name]
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        # Total/Summary Row for the table
        last_row = start_row + 2 + len(df)
        ws.cell(row=last_row, column=start_col, value="Total").font = Font(name="Segoe UI", size=9, bold=True)
        ws.cell(row=last_row, column=start_col).border = double_bottom_border
        
        for col_offset in range(1, len(headers)):
            c_idx = start_col + col_offset
            col_name = headers[col_offset]
            cell = ws.cell(row=last_row, column=c_idx)
            cell.border = double_bottom_border
            
            # Sum formula for numerical columns
            if col_name in ['Revenue', 'Quantity_Sold', 'Orders', 'Quantity Sold']:
                col_letter = get_column_letter(c_idx)
                start_cell = f"{col_letter}{start_row+2}"
                end_cell = f"{col_letter}{last_row-1}"
                cell.value = f"=SUM({start_cell}:{end_cell})"
                cell.font = Font(name="Segoe UI", size=9, bold=True)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col_name in number_formats:
                    cell.number_format = number_formats[col_name]
            elif col_name == 'Revenue Share %':
                # Sum formula for percentage
                col_letter = get_column_letter(c_idx)
                start_cell = f"{col_letter}{start_row+2}"
                end_cell = f"{col_letter}{last_row-1}"
                cell.value = f"=SUM({start_cell}:{end_cell})"
                cell.font = Font(name="Segoe UI", size=9, bold=True)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '0.00"%"'

    # 3. Write Tables side-by-side / staggered
    # Table 1: Top 10 Products (Column B)
    product_formats = {'Revenue': '$#,##0.00', 'Quantity_Sold': '#,##0', 'Orders': '#,##0'}
    write_table(start_row=7, start_col=2, title="Top 10 Products by Revenue", df=analysis['top_products'], number_formats=product_formats)
    
    # Table 2: Monthly Trends (Column G)
    trend_formats = {'Revenue': '$#,##0.00', 'Orders': '#,##0', 'Quantity_Sold': '#,##0'}
    write_table(start_row=7, start_col=7, title="Monthly Revenue Trend", df=analysis['monthly_trends'], number_formats=trend_formats)
    
    # Table 3: Regional Sales (Column L)
    region_formats = {'Revenue': '$#,##0.00', 'Orders': '#,##0', 'Revenue Share %': '0.00"%"'}
    write_table(start_row=7, start_col=12, title="Regional Sales Breakdown", df=analysis['regional_sales'], number_formats=region_formats)
    
    # Set custom column dimensions for spacing
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['F'].width = 3
    ws.column_dimensions['K'].width = 3
    
    # Auto-fit table columns
    table_cols = [
        (2, 4),   # B to D
        (7, 9),   # G to I
        (12, 14)  # L to N
    ]
    for start, end in table_cols:
        for col_idx in range(start, end + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(7, 20):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def generate_charts_images(analysis: dict, output_dir: str) -> dict:
    """
    Generates beautiful charts using Matplotlib and saves them to the output dir.
    """
    charts_dir = os.path.join(output_dir, "charts")
    if not os.path.exists(charts_dir):
        os.makedirs(charts_dir)
        
    # Styles config
    plt.style.use('seaborn-v0_8-whitegrid' if 'seaborn-v0_8-whitegrid' in plt.style.available else 'default')
    primary_color = "#1F4E78"
    accent_colors = ["#1F4E78", "#2C5282", "#3182CE", "#4299E1", "#63B3ED", "#90CDF4", "#CBD5E0", "#E2E8F0"]
    
    # Chart 1: Top Products (take top 5 for bar chart)
    fig, ax = plt.subplots(figsize=(6, 4))
    top_5_products = analysis['top_products'].head(5)
    bars = ax.bar(top_5_products['Product'], top_5_products['Revenue'] / 1000, color=accent_colors[:5], edgecolor='none', width=0.6)
    ax.set_title("Top 5 Products by Revenue", fontsize=12, fontweight='bold', color='#1F4E78', pad=15)
    ax.set_ylabel("Revenue ($k)", fontsize=10, color='#4A5568')
    ax.set_xlabel("Product", fontsize=10, color='#4A5568')
    ax.tick_params(colors='#4A5568', labelsize=9)
    ax.grid(axis='y', linestyle='--', alpha=0.7)
    # Add values on top of bars
    for bar in bars:
        yval = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2.0, yval + (max(top_5_products['Revenue'])/10000), f"${yval:,.1f}k", ha='center', va='bottom', fontsize=8, fontweight='bold', color='#2D3748')
    plt.tight_layout()
    chart1_path = os.path.join(charts_dir, "top_products.png")
    fig.savefig(chart1_path, dpi=150)
    plt.close(fig)
    
    # Chart 2: Monthly Revenue Trend
    fig, ax = plt.subplots(figsize=(6, 4))
    trends = analysis['monthly_trends']
    ax.plot(trends['Month'], trends['Revenue'] / 1000, marker='o', linewidth=2.5, color=primary_color, markerfacecolor='#3182CE', markersize=6)
    ax.set_title("Monthly Revenue Trend", fontsize=12, fontweight='bold', color='#1F4E78', pad=15)
    ax.set_ylabel("Revenue ($k)", fontsize=10, color='#4A5568')
    ax.set_xlabel("Month", fontsize=10, color='#4A5568')
    ax.tick_params(colors='#4A5568', labelsize=9)
    ax.grid(True, linestyle='--', alpha=0.7)
    # Add grid and values
    for x, y in zip(trends['Month'], trends['Revenue'] / 1000):
        ax.text(x, y + (max(trends['Revenue'])/100000), f"${y:,.1f}k", ha='center', va='bottom', fontsize=8, color='#2D3748')
    plt.tight_layout()
    chart2_path = os.path.join(charts_dir, "monthly_trend.png")
    fig.savefig(chart2_path, dpi=150)
    plt.close(fig)
    
    # Chart 3: Regional Share Donut Chart
    fig, ax = plt.subplots(figsize=(5.5, 4))
    regions = analysis['regional_sales']
    wedges, texts, autotexts = ax.pie(
        regions['Revenue'], 
        labels=regions['Region'], 
        autopct='%1.1f%%', 
        startangle=90, 
        colors=accent_colors[:len(regions)], 
        textprops=dict(color='#2D3748', size=9),
        wedgeprops=dict(width=0.4, edgecolor='w', linewidth=2) # Donut style
    )
    # Style percentages inside donut
    for autotext in autotexts:
        autotext.set_fontsize(8)
        autotext.set_weight('bold')
        autotext.set_color('white')
    ax.set_title("Revenue Share by Region", fontsize=12, fontweight='bold', color='#1F4E78', pad=15)
    plt.tight_layout()
    chart3_path = os.path.join(charts_dir, "regional_share.png")
    fig.savefig(chart3_path, dpi=150)
    plt.close(fig)
    
    return {
        'top_products': chart1_path,
        'monthly_trend': chart2_path,
        'regional_share': chart3_path
    }
