import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Headless matplotlib
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

def export_expenses_to_excel(expenses: list, output_path: str):
    """
    Creates a styled multi-sheet Excel report of expenses with category aggregates and embedded charts in Indian Rupees (INR).
    """
    # 1. Prepare Data
    if not expenses:
        # Create a dummy row so openpyxl doesn't crash on empty lists
        expenses = [{'id': 0, 'amount': 0.0, 'category': 'None', 'description': 'No expenses recorded', 'date': '2026-06-14'}]
        
    df = pd.DataFrame(expenses)
    df = df[['date', 'category', 'amount', 'description']]
    df.columns = ['Date', 'Category', 'Amount', 'Description']
    df['Date'] = pd.to_datetime(df['Date'])
    df = df.sort_values(by='Date') # Chronological order
    
    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    # Create Workbook
    wb = openpyxl.Workbook()
    ws_list = wb.active
    ws_list.title = "Expense Log"
    
    # Styles config
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid") # Finance Green
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    double_bottom_border = Border(
        bottom=Side(style='double', color='1B5E20'),
        top=Side(style='thin', color='D3D3D3')
    )
    
    # 2. Write Sheet 1: Expense Log
    ws_list.views.sheetView[0].showGridLines = True
    headers = list(df.columns)
    ws_list.append(headers)
    
    for row in dataframe_rows_formatted(df):
        ws_list.append(row)
        
    # Style Header
    for col_idx in range(1, len(headers) + 1):
        cell = ws_list.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Style Data
    for r_idx in range(2, ws_list.max_row + 1):
        for c_idx, col_name in enumerate(headers, 1):
            cell = ws_list.cell(row=r_idx, column=c_idx)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            
            # Format alignment & numeric representation
            if col_name == 'Amount':
                cell.alignment = right_align
                # Standard Indian formatting representation: ₹#,##,##0.00
                cell.number_format = '"₹"#,##,##0.00'
            elif col_name == 'Date':
                cell.alignment = center_align
                cell.number_format = 'YYYY-MM-DD'
            elif col_name == 'Category':
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Add Total Row in Sheet 1
    total_row_idx = ws_list.max_row + 1
    ws_list.cell(row=total_row_idx, column=1, value="Total").font = Font(name="Segoe UI", size=10, bold=True)
    ws_list.cell(row=total_row_idx, column=1).border = double_bottom_border
    
    for c_idx in range(2, len(headers) + 1):
        cell = ws_list.cell(row=total_row_idx, column=c_idx)
        cell.border = double_bottom_border
        if headers[c_idx-1] == 'Amount':
            col_letter = get_column_letter(c_idx)
            cell.value = f"=SUM({col_letter}2:{col_letter}{total_row_idx-1})"
            cell.font = Font(name="Segoe UI", size=10, bold=True)
            cell.alignment = right_align
            cell.number_format = '"₹"#,##,##0.00'
            
    # Auto-fit columns
    auto_fit_column_widths(ws_list)

    # 3. Write Sheet 2: Category Breakdown
    ws_breakdown = wb.create_sheet(title="Category Breakdown")
    ws_breakdown.views.sheetView[0].showGridLines = True
    
    # Calculate breakdown
    category_summary = df.groupby('Category').agg(Total_Spent=('Amount', 'sum')).sort_values(by='Total_Spent', ascending=False).reset_index()
    total_expense = df['Amount'].sum()
    category_summary['Share %'] = (category_summary['Total_Spent'] / total_expense * 100).round(2) if total_expense > 0 else 0.0
    
    ws_breakdown.append(["Category", "Total Spent (INR)", "Share %"])
    for row in dataframe_rows_formatted(category_summary):
        ws_breakdown.append(row)
        
    # Style Header
    for col_idx in range(1, 4):
        cell = ws_breakdown.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Style Data
    for r_idx in range(2, ws_breakdown.max_row + 1):
        for c_idx in range(1, 4):
            cell = ws_breakdown.cell(row=r_idx, column=c_idx)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            if c_idx == 1:
                cell.alignment = left_align
            elif c_idx == 2:
                cell.alignment = right_align
                cell.number_format = '"₹"#,##,##0.00'
            elif c_idx == 3:
                cell.alignment = right_align
                cell.number_format = '0.00"%"'
                
    # Add Total Row
    brk_total_row = ws_breakdown.max_row + 1
    ws_breakdown.cell(row=brk_total_row, column=1, value="Total").font = Font(name="Segoe UI", size=10, bold=True)
    ws_breakdown.cell(row=brk_total_row, column=1).border = double_bottom_border
    
    # Sum Total Spent
    cell_sum = ws_breakdown.cell(row=brk_total_row, column=2, value=f"=SUM(B2:B{brk_total_row-1})")
    cell_sum.font = Font(name="Segoe UI", size=10, bold=True)
    cell_sum.alignment = right_align
    cell_sum.number_format = '"₹"#,##,##0.00'
    cell_sum.border = double_bottom_border
    
    # Sum Share %
    cell_share = ws_breakdown.cell(row=brk_total_row, column=3, value=f"=SUM(C2:C{brk_total_row-1})")
    cell_share.font = Font(name="Segoe UI", size=10, bold=True)
    cell_share.alignment = right_align
    cell_share.number_format = '0.00"%"'
    cell_share.border = double_bottom_border
    
    auto_fit_column_widths(ws_breakdown)

    # 4. Generate Visual Charts
    charts_paths = generate_expense_charts(df, category_summary, output_dir)
    
    # 5. Sheet 3: Dashboard
    ws_dashboard = wb.create_sheet(title="Dashboard")
    ws_dashboard.views.sheetView[0].showGridLines = True
    
    # Title
    ws_dashboard.cell(row=2, column=2, value="Expense Analysis Dashboard (INR)").font = Font(name="Segoe UI", size=18, bold=True, color="2E7D32")
    
    # Embed Images
    from openpyxl.drawing.image import Image
    ws_dashboard.add_image(Image(charts_paths['category_pie']), 'B4')
    ws_dashboard.add_image(Image(charts_paths['trend_line']), 'L4')
    
    wb.save(output_path)
    return output_path

def dataframe_rows_formatted(df):
    """Generator to yield formatted row values suitable for openpyxl appending."""
    for row in df.itertuples(index=False):
        row_vals = []
        for v in row:
            if isinstance(v, pd.Timestamp):
                row_vals.append(v.strftime('%Y-%m-%d'))
            elif pd.isna(v):
                row_vals.append("")
            else:
                row_vals.append(v)
        yield row_vals

def auto_fit_column_widths(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is not None:
                val_str = str(val)
                if isinstance(val, float) and cell.number_format == '"₹"#,##,##0.00':
                    val_str = f"Rs.{val:,.2f}"
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def generate_expense_charts(df, category_summary, output_dir) -> dict:
    temp_dir = os.path.join(output_dir, "temp_charts")
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
        
    plt.style.use('seaborn-v0_8-whitegrid' if 'seaborn-v0_8-whitegrid' in plt.style.available else 'default')
    colors = ["#2E7D32", "#1976D2", "#E64A19", "#7B1FA2", "#FBC02D", "#C2185B", "#0097A7", "#5D4037"]
    
    # Chart 1: Category share
    fig, ax = plt.subplots(figsize=(6, 4))
    wedges, texts, autotexts = ax.pie(
        category_summary['Total_Spent'],
        labels=category_summary['Category'],
        autopct='%1.1f%%',
        colors=colors[:len(category_summary)],
        startangle=90,
        textprops=dict(color='#2D3748', size=9),
        wedgeprops=dict(width=0.4, edgecolor='w', linewidth=1.5)
    )
    for autotext in autotexts:
        autotext.set_fontsize(8)
        autotext.set_weight('bold')
        autotext.set_color('white')
    ax.set_title("Spending Share by Category", fontsize=12, fontweight='bold', color='#1B5E20', pad=15)
    plt.tight_layout()
    pie_path = os.path.join(temp_dir, "exp_category.png")
    fig.savefig(pie_path, dpi=150)
    plt.close(fig)
    
    # Chart 2: Daily Spending Trend
    fig, ax = plt.subplots(figsize=(7, 4))
    df_trend = df.groupby('Date').agg(Daily_Total=('Amount', 'sum')).reset_index()
    # Take last 30 days of data or sort chronological
    df_trend = df_trend.sort_values(by='Date')
    
    # Plot line
    ax.plot(df_trend['Date'], df_trend['Daily_Total'], marker='o', color='#2E7D32', linewidth=2, markersize=5, markerfacecolor='#1976D2')
    ax.set_title("Daily Spending Trend (INR)", fontsize=12, fontweight='bold', color='#1B5E20', pad=15)
    ax.set_ylabel("Spent (₹)", fontsize=10)
    ax.set_xlabel("Date", fontsize=10)
    
    # Format x-ticks
    fig.autofmt_xdate()
    ax.grid(True, linestyle='--', alpha=0.6)
    
    plt.tight_layout()
    trend_path = os.path.join(temp_dir, "exp_trend.png")
    fig.savefig(trend_path, dpi=150)
    plt.close(fig)
    
    return {
        'category_pie': pie_path,
        'trend_line': trend_path
    }
